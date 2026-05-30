#!/usr/bin/env python3
"""Extract product images from JCK Price List workbooks into images.js."""

from __future__ import annotations

import base64
import io
import json
import re
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

ROOT = Path(__file__).resolve().parents[1]
EXCEL_DIR = Path("/Users/rj/Downloads/JCK PRICE LIST")
SOURCES = [
    EXCEL_DIR / "JCK Price List-1.xlsx",
    EXCEL_DIR / "JCK Price List-2.xlsx",
    EXCEL_DIR / "JCK Price List-3.xlsx",
]
OUT = ROOT / "images.js"
HTML_FILES = [ROOT / "index.html", ROOT / "lookup.html"]
MAX_PX = 360
JPEG_QUALITY = 74


def code_for_row(code_by_row: dict[int, str], row: int) -> str | None:
    code = code_by_row.get(row)
    if code:
        return code
    for rr in range(row, 1, -1):
        if rr in code_by_row:
            return code_by_row[rr]
    return None


def compress_image(data: bytes) -> str:
    im = Image.open(io.BytesIO(data))
    if im.mode not in ("RGB", "L"):
        im = im.convert("RGB")
    im.thumbnail((MAX_PX, MAX_PX), Image.Resampling.LANCZOS)
    buf = io.BytesIO()
    im.save(buf, format="JPEG", quality=JPEG_QUALITY, optimize=True)
    b64 = base64.b64encode(buf.getvalue()).decode("ascii")
    return f"data:image/jpeg;base64,{b64}"


def extract_workbook(path: Path) -> dict[str, str]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    code_by_row: dict[int, str] = {}
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val:
            code_by_row[r] = str(val).strip().upper()

    images: dict[str, str] = {}
    for img in ws._images:
        row = img.anchor._from.row + 1
        code = code_for_row(code_by_row, row)
        if not code:
            continue
        try:
            images[code] = compress_image(img._data())
        except Exception as err:
            print(f"Warning: skip {code} in {path.name}: {err}")
    return images


def merge_sources(paths: list[Path]) -> tuple[dict[str, str], dict[str, int]]:
    merged: dict[str, str] = {}
    per_file: dict[str, int] = {}
    for path in paths:
        if not path.exists():
            print(f"Warning: missing {path}")
            continue
        items = extract_workbook(path)
        per_file[path.name] = len(items)
        merged.update(items)
    return merged, per_file


def patch_html_remove_inline_imgs(html_path: Path):
    text = html_path.read_text(encoding="utf-8")
    marker = "const IMGS="
    idx = text.find(marker)
    if idx < 0:
        return
    obj_start = text.index("{", idx)
    depth = 0
    end = None
    for i in range(obj_start, len(text)):
        if text[i] == "{":
            depth += 1
        elif text[i] == "}":
            depth -= 1
            if depth == 0:
                end = i
                break
    if end is None:
        raise SystemExit(f"Could not parse IMGS object in {html_path}")
    text = text[:idx] + text[end + 1 :]
    script_tag = '<script src="images.js?v=20260529b"></script>'
    if script_tag not in text:
        cfg = '<script src="cloud-config.js?v=20260529a"></script>'
        if cfg in text:
            text = text.replace(cfg, cfg + "\n" + script_tag, 1)
        else:
            raise SystemExit(f"cloud-config script tag not found in {html_path}")
    html_path.write_text(text, encoding="utf-8")


def main():
    missing = [p for p in SOURCES if not p.exists()]
    if len(missing) == len(SOURCES):
        raise SystemExit(f"No workbooks found under {EXCEL_DIR}")

    images, per_file = merge_sources(SOURCES)
    if not images:
        raise SystemExit("No images extracted")

    payload = json.dumps(images, separators=(",", ":"))
    OUT.write_text(f"const IMGS={payload};\n", encoding="utf-8")

    for html in HTML_FILES:
        if html.exists():
            patch_html_remove_inline_imgs(html)

    print(f"Wrote {len(images)} images to {OUT} ({OUT.stat().st_size // 1024 // 1024} MB)")
    for name, count in per_file.items():
        print(f"  {name}: {count} images")
    for html in HTML_FILES:
        if html.exists():
            print(f"Patched {html.name} to load images.js")


if __name__ == "__main__":
    main()
