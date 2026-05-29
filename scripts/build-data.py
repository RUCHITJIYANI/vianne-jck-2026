#!/usr/bin/env python3
"""Regenerate ITEMS_RAW from JCK Price List workbooks (1, 2, 3)."""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
EXCEL_DIR = Path("/Users/rj/Downloads/JCK PRICE LIST")
SOURCES = [
    EXCEL_DIR / "JCK Price List-1.xlsx",
    EXCEL_DIR / "JCK Price List-2.xlsx",
    EXCEL_DIR / "JCK Price List-3.xlsx",
]
OUT = ROOT / "data.js"
HTML_FILES = [ROOT / "index.html", ROOT / "lookup.html"]

MARGIN_BY_COLLECTION = {
    "CLASSICS": 0.85,
    "ROSE": 0.75,
    "LINQ": 0.65,
    "BEZEL": 0.75,
    "OTHER": 0.75,
}


def num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def int_or_none(v):
    n = num(v)
    return int(n) if n is not None else None


def clean_str(v):
    if v is None:
        return ""
    return str(v).strip()


def margin_for(collection: str) -> float:
    key = clean_str(collection).upper()
    for name, m in MARGIN_BY_COLLECTION.items():
        if name in key:
            return m
    return 0.75


def parse_workbook(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    items = []
    current = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[0]
        if code:
            if current:
                items.append(current)
            coll = clean_str(row[3])
            today_tariff = num(row[20]) if len(row) > 20 else None
            margin = margin_for(coll)
            sale_calc = round(today_tariff / margin, 2) if today_tariff and margin else None
            current = {
                "unique_code": clean_str(code),
                "style_code": clean_str(row[1]),
                "design": clean_str(row[2]),
                "collection": coll,
                "size": clean_str(row[4]) if row[4] else "",
                "qty": int_or_none(row[5]) or 0,
                "kt_col": clean_str(row[6]),
                "gross_wt": num(row[8]) if len(row) > 8 else None,
                "net_wt": num(row[9]) if len(row) > 9 else None,
                "stones": [],
                "today_cost": num(row[17]) if len(row) > 17 else None,
                "inward_value": num(row[18]) if len(row) > 18 else None,
                "today_cost_tariff": today_tariff,
                "inward_tariff": num(row[21]) if len(row) > 21 else None,
                "margin": margin,
                "sale_price_calc": sale_calc,
                "round_off": num(row[24]) if len(row) > 24 else None,
            }

        if not current:
            continue

        shape = row[10] if len(row) > 10 else None
        if shape:
            current["stones"].append(
                {
                    "shape": clean_str(shape),
                    "clarity": clean_str(row[11]) if len(row) > 11 else "",
                    "pcs": int_or_none(row[12]) if len(row) > 12 else None,
                    "total_pcs": int_or_none(row[13]) if len(row) > 13 else None,
                    "cts": num(row[14]) if len(row) > 14 else None,
                    "total_cts": num(row[15]) if len(row) > 15 else None,
                }
            )

    if current:
        items.append(current)

    return items


def merge_sources(paths):
    """Later workbooks override earlier entries for the same unique_code."""
    by_code = {}
    per_file = {}
    for path in paths:
        if not path.exists():
            print(f"Warning: missing {path}")
            continue
        items = parse_workbook(path)
        per_file[path.name] = len(items)
        for item in items:
            by_code[item["unique_code"].upper()] = item
    merged = sorted(by_code.values(), key=lambda x: x["unique_code"])
    return merged, per_file


def patch_html(html_path: Path, items_json: str, count: int):
    text = html_path.read_text(encoding="utf-8")
    marker = "const ITEMS_RAW="
    idx = text.find(marker)
    if idx < 0:
        raise SystemExit(f"ITEMS_RAW not found in {html_path}")
    arr_start = text.index("[", idx)
    depth = 0
    arr_end = None
    for i in range(arr_start, len(text)):
        if text[i] == "[":
            depth += 1
        elif text[i] == "]":
            depth -= 1
            if depth == 0:
                arr_end = i
                break
    if arr_end is None:
        raise SystemExit(f"Could not parse ITEMS_RAW array in {html_path}")
    text = text[:idx] + f"const ITEMS_RAW={items_json}" + text[arr_end + 1 :]
    text = re.sub(
        r'(<span id="hdrItemCount">)\d+(</span> ITEMS)',
        rf"\g<1>{count}\g<2>",
        text,
    )
    html_path.write_text(text, encoding="utf-8")


def main():
    missing = [p for p in SOURCES if not p.exists()]
    if len(missing) == len(SOURCES):
        raise SystemExit(f"No workbooks found under {EXCEL_DIR}")

    items, per_file = merge_sources(SOURCES)
    if not items:
        raise SystemExit("No items parsed from workbooks")

    items_json = json.dumps(items, separators=(",", ":"))

    OUT.write_text(
        "const ITEMS_RAW="
        + items_json
        + ";\nconst ITEMS={};\nITEMS_RAW.forEach(i=>{ITEMS[i.unique_code.toUpperCase()]=i;});\n",
        encoding="utf-8",
    )

    for html in HTML_FILES:
        if html.exists():
            patch_html(html, items_json, len(items))

    print(f"Merged {len(items)} unique items into {OUT}")
    for name, count in per_file.items():
        print(f"  {name}: {count} rows parsed")
    for html in HTML_FILES:
        if html.exists():
            print(f"Patched {html.name}")


if __name__ == "__main__":
    main()
